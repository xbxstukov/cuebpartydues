from PyQt5 import QtWidgets
from Ui_party_dues_gui_hd import Ui_MainWindow
import sys
import os
import datetime
import sys
import openpyxl
import time
import win32com.client as win32
from openpyxl.styles import Font, colors, Border, Side, Alignment, PatternFill

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *

class mywindow(QtWidgets.QMainWindow, Ui_MainWindow):
    
    def __init__(self):
        super(mywindow,self).__init__()
        self.setupUi(self)
        self.setFixedSize(self.width(), self.height())
        self.importButton.clicked.connect(self.fileSelect)
        self.actionImport.triggered.connect(self.fileSelect)
        self.calculateButton.clicked.connect(self.calculation)

    def fileSelect(self):
        filename, filetype = QtWidgets.QFileDialog.getOpenFileName(self,'选择工资文件','','Excel 2003文件(*.xls)')
        print(filename)
        filepathDisplay = str(filename).replace("/","\\")
        self.lineEdit.setText(filepathDisplay)
        global filepath 
        filepath = str(filename).replace("/","\\")
        self.calculateButton.setEnabled(True)

    def calculation(self):
        unitName, okPressed = QtWidgets.QInputDialog.getText(self, "单位名称", "请输入您的单位名称", QtWidgets.QLineEdit.Normal, "工商管理学院")
        # 获取当前年份，判定季度 
        year = datetime.datetime.now().strftime('%Y')
        # 获取当前月份，判定季度
        month = datetime.datetime.now().strftime('%m')
        if int(month) < 3:
            # print('\n还没有到3月，无法计算第一季度党费，程序将于5秒后自动退出...')
            QtWidgets.QMessageBox.information(self, "提示", "还没有到3月，无法计算第一季度党费，程序即将关闭！", QtWidgets.QMessageBox.Yes)
            exit()
        elif int(month) >= 3 and int(month) < 6:
            season = 1
            cellMonth1 = '1'
            cellMonth2 = '2'
            cellMonth3 = '3'
        elif int(month) >= 6 and int(month) < 9:
            season = 2
            cellMonth1 = '4'
            cellMonth2 = '5'
            cellMonth3 = '6'
        elif int(month) >= 9 and int(month) < 12:
            season = 3
            cellMonth1 = '7'
            cellMonth2 = '8'
            cellMonth3 = '9'
        else:
            season = 4
            cellMonth1 = '10'
            cellMonth2 = '11'
            cellMonth3 = '12'
        # time.sleep(2)
        salaryPath = filepath
        # 转换为xlsx文件
        # xlspath = os.path.abspath(salaryPath) 
        xlspath = salaryPath
        xlsxpath = xlspath + 'x'
        current_dir = os.path.dirname(os.path.abspath(xlsxpath))
        if os.path.exists(xlsxpath):
            os.remove(xlsxpath)
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(xlspath)
        wb.SaveAs(xlsxpath, FileFormat=51)    # FileFormat = 51 is for .xlsx extension
        wb.Close()                               # FileFormat = 56 is for .xls extension
        # 调整格式
        wbx = openpyxl.load_workbook(xlsxpath)
        wsx = wbx.active
        # 单元格文本转换为数字
        for i in range(2, wsx.max_row+1):
            if season == 1:
                January = str(wsx['B'+str(i)].value)
                January = January.replace(",","")
                wsx['B'+str(i)] = float(January)
                Feburary = str(wsx['C'+str(i)].value)
                Feburary = Feburary.replace(",","")
                wsx['C'+str(i)] = float(Feburary)
                March = str(wsx['D'+str(i)].value)
                March = March.replace(",","")
                wsx['D'+str(i)] = float(March)
            elif season == 2:
                April = str(wsx['E'+str(i)].value)
                April = April.replace(",","")
                wsx['E'+str(i)] = float(April)
                May = str(wsx['F'+str(i)].value)
                May = May.replace(",","")
                wsx['F'+str(i)] = float(May)
                June = str(wsx['G'+str(i)].value)
                June = June.replace(",","")
                wsx['G'+str(i)] = float(June)
            elif season == 3:
                July = str(wsx['H'+str(i)].value)
                July = July.replace(",","")
                wsx['H'+str(i)] = float(July)
                August = str(wsx['I'+str(i)].value)
                August = August.replace(",","")   
                wsx['I'+str(i)] = float(August)
                September = str(wsx['J'+str(i)].value)
                September = September.replace(",","")
                wsx['J'+str(i)] = float(September)
            elif season == 4:
                October = str(wsx['K'+str(i)].value)
                October = October.replace(",","")
                wsx['K'+str(i)] = float(October)
                November = str(wsx['L'+str(i)].value)
                November = November.replace(",","")
                wsx['L'+str(i)] = float(November)
                December = str(wsx['M'+str(i)].value)
                December = December.replace(",","")
                wsx['M'+str(i)] = float(December)
        wbx.save(xlsxpath)
        # print()
        # print('正在转换Excel文件: 工资.xls → 工资.xlsx')
        salary = openpyxl.load_workbook(xlsxpath)
        sheet = salary.active
        dueNameList = []
        JanList = []
        FebList = []
        MarList = []
        AprList = []
        MayList = []
        JunList = []
        JulList = []
        AugList = []
        SepList = []
        OctList = []
        NovList = []
        DecList = []
        for row in range(2, sheet.max_row+1):
            dueName = sheet['A'+str(row)].value
            dueNameList.append(dueName)
            Jan = sheet['B'+str(row)].value
            JanList.append(Jan)
            Feb = sheet['C'+str(row)].value
            FebList.append(Feb)
            Mar = sheet['D'+str(row)].value
            MarList.append(Mar)
            Apr = sheet['E'+str(row)].value
            AprList.append(Apr)
            May = sheet['F'+str(row)].value
            MayList.append(May)
            Jun = sheet['G'+str(row)].value
            JunList.append(Jun)
            Jul = sheet['H'+str(row)].value
            JulList.append(Jul)
            Aug = sheet['I'+str(row)].value
            AugList.append(Aug)
            Sep = sheet['J'+str(row)].value
            SepList.append(Sep)
            Oct = sheet['K'+str(row)].value
            OctList.append(Oct)
            Nov = sheet['L'+str(row)].value
            NovList.append(Nov)
            Dec = sheet['M'+str(row)].value
            DecList.append(Dec)

        gwgzDex = dueNameList.index('岗位工资')
        xjgzDex = dueNameList.index('薪级工资')
        zbDex = dueNameList.index('职补')
        shbfDex = dueNameList.index('书报费')
        xlfDex = dueNameList.index('洗理费')
        shfDex = dueNameList.index('生活费')
        xnbtDex = dueNameList.index('校内补贴')
        try:
            gtDex = dueNameList.index('岗贴')
            isFindGt = True
        except ValueError:
            isFindGt = False
            # print('\n无法在您提供的工资表中找到“岗贴”，无法自动生成缴费表')
            QtWidgets.QMessageBox.information(self, "提示", "无法在您提供的工资表中找到“岗贴”，请手动输入岗贴", QtWidgets.QMessageBox.Yes)
            bzgt, okPressed = QtWidgets.QInputDialog.getInt(self, "自定义岗贴", "请输入岗贴(整数):", 0, 5400, 14100, 100)
            # # time.sleep(2)
            # while True:
            #     bzgt = input('\n请您手动输入标准岗贴，然后按回车键：')
            #     if bzgt.isdigit():
            #         break
            #     else:
            #         print('\n您输入的不是整数，请输入整数')
        jsheDex = dueNameList.index('绩效减少额')
        kgjjDex = dueNameList.index('扣公积金')
        kzhynjDex = dueNameList.index('扣职业年金')
        kyxDex = dueNameList.index('扣养险(新老合并）')
        kshxDex = dueNameList.index('扣失险')
        kylxDex = dueNameList.index('扣医疗险')
        try:
            srzj3Dex = dueNameList.index('薪金收入合计(3项)')
        except ValueError:
            # print('\n无法在您提供的工资表中找到“薪金收入合计(3项)”项目，请您自行修改此项目在工资表中的名称，或联系作者索要最新版本的程序！')
            # time.sleep(2)
            # print('\n程序将于5秒后自动退出...')
            # time.sleep(5)
            QtWidgets.QMessageBox.information(self, "错误提示", "无法在您提供的工资表中找到“薪金收入合计(3项)”项目，请您自行修改此项目在工资表中的名称，程序将关闭", QtWidgets.QMessageBox.Yes)
            exit()
        ksh2Dex = dueNameList.index('扣税2') 

        # time.sleep(2)
        # print()
        # print('正在解析工资数据...')
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        # 缴费模板生成
        for j in range(1,10):
            ws2.row_dimensions[j].height = 30    # 1-9行 行高30
        ws2.column_dimensions['A'].width = 7.28
        # 列宽B-P
        ws2.column_dimensions['B'].width = 7.02
        ws2.column_dimensions['C'].width = 7.02
        ws2.column_dimensions['D'].width = 7.02
        ws2.column_dimensions['E'].width = 7.02
        ws2.column_dimensions['F'].width = 7.02
        ws2.column_dimensions['G'].width = 7.02
        ws2.column_dimensions['H'].width = 7.02
        ws2.column_dimensions['I'].width = 7.02
        ws2.column_dimensions['J'].width = 7.02
        ws2.column_dimensions['K'].width = 7.02
        ws2.column_dimensions['L'].width = 7.02
        ws2.column_dimensions['M'].width = 7.02
        ws2.column_dimensions['N'].width = 7.02
        ws2.column_dimensions['O'].width = 7.02
        ws2.column_dimensions['P'].width = 7.02
        # 列宽Q，R
        ws2.column_dimensions['Q'].width = 9.17
        ws2.column_dimensions['R'].width = 9.17
        # 合并A1-R1,作为标题
        ws2.merge_cells('A1:R1')
        # titleCell = ws2.cell(row = 1, column = 1)
        titleCell = ws2['A1']
        titleCell.alignment = Alignment(horizontal= 'center', vertical = 'center')
        titleCell.font = Font(size=16, bold= True)
        ws2['A1'] = unitName + '教师党员应交党费统计表（' + str(year) + '年第'+str(season)+'季度）'
        # A2 年 月
        ws2['A2'] = '年 月'
        for n in range(1,19):
            # 遍历第2行的标题格式
            subTitleCell = ws2.cell(column = n, row = 2)
            subTitleCell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            subTitleCell.fill = PatternFill(fill_type='solid',start_color='FF8FBC8F',end_color='FF8FBC8F')
            subTitleCell.border = Border(left=Side(border_style='thin',color='FF000000'),
            right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='thin',color='FF000000'),
            bottom=Side(border_style='thin',color='FF000000')
            )
            # 遍历第9行的格式
            bottomCell = ws2.cell(column = n, row = 9)
            bottomCell.alignment = Alignment(horizontal='left', vertical='center')
            bottomCell.border = Border(left=Side(border_style='thin',color='FF000000'),
            right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='thin',color='FF000000'),
            bottom=Side(border_style='thin',color='FF000000')
            )
            for m in range(3,9):
                # 遍历第3、4、5行的数据格式
                contentCell = ws2.cell(column = n, row = m)
                contentCell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                contentCell.border = Border(left=Side(border_style='thin',color='FF000000'),
                right=Side(border_style='thin',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='thin',color='FF000000')
                )
        ws2['A3'] = str(year) + '年' + cellMonth1 + '月'
        ws2['A4'] = str(year) + '年' + cellMonth2 + '月'
        ws2['A5'] = str(year) + '年' + cellMonth3 + '月'
        ws2['A6'] = '总计'
        # 合并B6-R1,作为总计数据
        ws2.merge_cells('B6:R6')
        # 合并A7-R7，作为公式展示
        ws2.merge_cells('A7:R7')
        ws2['A7'] = '所得税计算公式：所得税=（岗位工资+薪级工资+职务补贴+书报费+洗理费+生活费+校内补贴+岗位津贴）/薪金收入合计(3项)*扣税2'
        # 合并A8:B8，作为党费收缴标准
        ws2.merge_cells('A8:B8')
        # 合并C8：R8，作为标准展示
        ws2.merge_cells('C8:R8')
        ws2['A8'] = '党费收交标准：'
        ws2['C8'] = '3000元（含）以下：0.5%，3000-5000元（含）：1%，5000元-10000元（含）：1.5%，10000元以上：2%'
        # 合并A9：D9，作为姓名（签字）
        ws2.merge_cells('A9:D9')
        ws2['A9'] = '姓名（签字）：'
        # 合并E9：H9，作为应交总金额
        ws2.merge_cells('E9:H9')
        # 合并I9：M9，作为所在支部
        ws2.merge_cells('I9:M9')
        ws2['I9'] = '所在支部：'
        # 合并N9：R9，作为支部书记（签字）
        ws2.merge_cells('N9:R9')
        ws2['N9'] = '支部书记（签字）：'
        ws2['B2'] = '岗位工资'
        ws2['C2'] = '+薪级工资'
        ws2['D2'] = '+职务补贴'
        ws2['E2'] = '+书报费'
        ws2['F2'] = '+洗理费'
        ws2['G2'] = '+生活费'
        ws2['H2'] = '+校内补贴'
        ws2['I2'] = '+岗位津贴'
        ws2['J2'] = '-减少额'
        ws2['K2'] = '-扣公积金'
        ws2['L2'] = '-扣职业年金'
        ws2['M2'] = '-扣养险'
        ws2['N2'] = '-扣失险'
        ws2['O2'] = '-扣医疗险'
        ws2['P2'] = '-所得税'
        ws2['Q2'] = '合计'
        ws2['R2'] = '应交党费'
        # 一季度
        if season == 1:
            # 第3行数据
            ws2['B3'] = round(JanList[gwgzDex], 2)
            ws2['C3'] = round(JanList[xjgzDex], 2)
            ws2['D3'] = round(JanList[zbDex], 2)
            ws2['E3'] = round(JanList[shbfDex], 2)
            ws2['F3'] = round(JanList[xlfDex], 2)
            ws2['G3'] = round(JanList[shfDex], 2)
            ws2['H3'] = round(JanList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I3'] = round(JanList[gtDex], 2)
            elif isFindGt is False:
                ws2['I3'] = round(float(bzgt), 2)
            ws2['J3'] = round(JanList[jsheDex], 2)
            ws2['K3'] = round(JanList[kgjjDex], 2)
            ws2['L3'] = round(JanList[kzhynjDex], 2)
            ws2['M3'] = round(JanList[kyxDex], 2)
            ws2['N3'] = round(JanList[kshxDex], 2)
            ws2['O3'] = round(JanList[kylxDex], 2)
            if isFindGt is True:
                tax = (JanList[gwgzDex]+JanList[xjgzDex]+JanList[zbDex]+JanList[shbfDex]+JanList[xlfDex]+JanList[shfDex]+\
                    JanList[xnbtDex]+JanList[gtDex])/JanList[srzj3Dex]*JanList[ksh2Dex]
            elif isFindGt is False:
                tax = (JanList[gwgzDex]+JanList[xjgzDex]+JanList[zbDex]+JanList[shbfDex]+JanList[xlfDex]+JanList[shfDex]+\
                    JanList[xnbtDex]+float(bzgt))/JanList[srzj3Dex]*JanList[ksh2Dex]        
            ws2['P3'] = round(tax, 2)
            if isFindGt is True:
                total = JanList[gwgzDex]+JanList[xjgzDex]+JanList[zbDex]+JanList[shbfDex]+JanList[xlfDex]+\
                    JanList[shfDex]+JanList[xnbtDex]+JanList[gtDex]-JanList[jsheDex]-JanList[kgjjDex]-JanList[kzhynjDex]-\
                        JanList[kyxDex]-JanList[kshxDex]-JanList[kylxDex]-tax
            elif isFindGt is False:
                total = JanList[gwgzDex]+JanList[xjgzDex]+JanList[zbDex]+JanList[shbfDex]+JanList[xlfDex]+\
                    JanList[shfDex]+JanList[xnbtDex]+float(bzgt)-JanList[jsheDex]-JanList[kgjjDex]-JanList[kzhynjDex]-\
                        JanList[kyxDex]-JanList[kshxDex]-JanList[kylxDex]-tax
            ws2['Q3'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R3'] = round(partyFee, 2)
            # ws2['S3'] = JanList[srzj3Dex]
            # ws2['T3'] = JanList[ksh2Dex]
            # 第4行数据
            ws2['B4'] = round(FebList[gwgzDex], 2)
            ws2['C4'] = round(FebList[xjgzDex], 2)
            ws2['D4'] = round(FebList[zbDex], 2)
            ws2['E4'] = round(FebList[shbfDex], 2)
            ws2['F4'] = round(FebList[xlfDex], 2)
            ws2['G4'] = round(FebList[shfDex], 2)
            ws2['H4'] = round(FebList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I4'] = round(FebList[gtDex], 2)
            elif isFindGt is False:
                ws2['I4'] = round(float(bzgt), 2)
            ws2['J4'] = round(FebList[jsheDex], 2)
            ws2['K4'] = round(FebList[kgjjDex], 2)
            ws2['L4'] = round(FebList[kzhynjDex], 2)
            ws2['M4'] = round(FebList[kyxDex], 2)
            ws2['N4'] = round(FebList[kshxDex], 2)
            ws2['O4'] = round(FebList[kylxDex], 2)
            if isFindGt is True:
                tax = (FebList[gwgzDex]+FebList[xjgzDex]+FebList[zbDex]+FebList[shbfDex]+FebList[xlfDex]+FebList[shfDex]+\
                    FebList[xnbtDex]+FebList[gtDex])/FebList[srzj3Dex]*FebList[ksh2Dex]
            elif isFindGt is False:
                tax = (FebList[gwgzDex]+FebList[xjgzDex]+FebList[zbDex]+FebList[shbfDex]+FebList[xlfDex]+FebList[shfDex]+\
                    FebList[xnbtDex]+float(bzgt))/FebList[srzj3Dex]*FebList[ksh2Dex]
            ws2['P4'] = round(tax, 2)
            if isFindGt is True:
                total = FebList[gwgzDex]+FebList[xjgzDex]+FebList[zbDex]+FebList[shbfDex]+FebList[xlfDex]+\
                    FebList[shfDex]+FebList[xnbtDex]+FebList[gtDex]-FebList[jsheDex]-FebList[kgjjDex]-FebList[kzhynjDex]-\
                        FebList[kyxDex]-FebList[kshxDex]-FebList[kylxDex]-tax
            elif isFindGt is False:
                total = FebList[gwgzDex]+FebList[xjgzDex]+FebList[zbDex]+FebList[shbfDex]+FebList[xlfDex]+\
                FebList[shfDex]+FebList[xnbtDex]+float(bzgt)-FebList[jsheDex]-FebList[kgjjDex]-FebList[kzhynjDex]-\
                    FebList[kyxDex]-FebList[kshxDex]-FebList[kylxDex]-tax
            ws2['Q4'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R4'] = round(partyFee, 2)
            # ws2['S4'] = FebList[srzj3Dex]
            # ws2['T4'] = FebList[ksh2Dex]
            # 第5行数据
            ws2['B5'] = round(MarList[gwgzDex], 2)
            ws2['C5'] = round(MarList[xjgzDex], 2)
            ws2['D5'] = round(MarList[zbDex], 2)
            ws2['E5'] = round(MarList[shbfDex], 2)
            ws2['F5'] = round(MarList[xlfDex], 2)
            ws2['G5'] = round(MarList[shfDex], 2)
            ws2['H5'] = round(MarList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I5'] = round(MarList[gtDex], 2)
            elif isFindGt is False:
                ws2['I5'] = round(float(bzgt), 2)
            ws2['J5'] = round(MarList[jsheDex], 2)
            ws2['K5'] = round(MarList[kgjjDex], 2)
            ws2['L5'] = round(MarList[kzhynjDex], 2)
            ws2['M5'] = round(MarList[kyxDex], 2)
            ws2['N5'] = round(MarList[kshxDex], 2)
            ws2['O5'] = round(MarList[kylxDex], 2)
            if isFindGt is True:
                tax = (MarList[gwgzDex]+MarList[xjgzDex]+MarList[zbDex]+MarList[shbfDex]+MarList[xlfDex]+MarList[shfDex]+\
                    MarList[xnbtDex]+MarList[gtDex])/MarList[srzj3Dex]*MarList[ksh2Dex]
            elif isFindGt is False:
                tax = (MarList[gwgzDex]+MarList[xjgzDex]+MarList[zbDex]+MarList[shbfDex]+MarList[xlfDex]+MarList[shfDex]+\
                    MarList[xnbtDex]+float(bzgt))/MarList[srzj3Dex]*MarList[ksh2Dex]
            ws2['P5'] = round(tax, 2)
            if isFindGt is True:
                total = MarList[gwgzDex]+MarList[xjgzDex]+MarList[zbDex]+MarList[shbfDex]+MarList[xlfDex]+\
                    MarList[shfDex]+MarList[xnbtDex]+MarList[gtDex]-MarList[jsheDex]-MarList[kgjjDex]-MarList[kzhynjDex]-\
                        MarList[kyxDex]-MarList[kshxDex]-MarList[kylxDex]-tax
            elif isFindGt is False:
                total = MarList[gwgzDex]+MarList[xjgzDex]+MarList[zbDex]+MarList[shbfDex]+MarList[xlfDex]+\
                    MarList[shfDex]+MarList[xnbtDex]+float(bzgt)-MarList[jsheDex]-MarList[kgjjDex]-MarList[kzhynjDex]-\
                        MarList[kyxDex]-MarList[kshxDex]-MarList[kylxDex]-tax
            ws2['Q5'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R5'] = round(partyFee, 2)
            # ws2['S5'] = MarList[srzj3Dex]
            # ws2['T5'] = MarList[ksh2Dex]
        # 二季度
        elif season == 2:
            # 第3行数据
            ws2['B3'] = round(AprList[gwgzDex], 2)
            ws2['C3'] = round(AprList[xjgzDex], 2)
            ws2['D3'] = round(AprList[zbDex], 2)
            ws2['E3'] = round(AprList[shbfDex], 2)
            ws2['F3'] = round(AprList[xlfDex], 2)
            ws2['G3'] = round(AprList[shfDex], 2)
            ws2['H3'] = round(AprList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I3'] = round(AprList[gtDex], 2)
            elif isFindGt is False:
                ws2['I3'] = round(float(bzgt), 2)
            ws2['J3'] = round(AprList[jsheDex], 2)
            ws2['K3'] = round(AprList[kgjjDex], 2)
            ws2['L3'] = round(AprList[kzhynjDex], 2)
            ws2['M3'] = round(AprList[kyxDex], 2)
            ws2['N3'] = round(AprList[kshxDex], 2)
            ws2['O3'] = round(AprList[kylxDex], 2)
            # ws2['S3'] = AprList[srzj3Dex]
            # ws2['T3'] = AprList[ksh2Dex]
            if isFindGt is True:
                tax = (AprList[gwgzDex]+AprList[xjgzDex]+AprList[zbDex]+AprList[shbfDex]+AprList[xlfDex]+AprList[shfDex]+\
                    AprList[xnbtDex]+AprList[gtDex])/AprList[srzj3Dex]*AprList[ksh2Dex]
            elif isFindGt is False:
                tax = (AprList[gwgzDex]+AprList[xjgzDex]+AprList[zbDex]+AprList[shbfDex]+AprList[xlfDex]+AprList[shfDex]+\
                    AprList[xnbtDex]+float(bzgt))/AprList[srzj3Dex]*AprList[ksh2Dex]
            ws2['P3'] = round(tax, 2)
            if isFindGt is True:
                total = AprList[gwgzDex]+AprList[xjgzDex]+AprList[zbDex]+AprList[shbfDex]+AprList[xlfDex]+\
                    AprList[shfDex]+AprList[xnbtDex]+AprList[gtDex]-AprList[jsheDex]-AprList[kgjjDex]-AprList[kzhynjDex]-\
                        AprList[kyxDex]-AprList[kshxDex]-AprList[kylxDex]-tax
            elif isFindGt is False:
                total = AprList[gwgzDex]+AprList[xjgzDex]+AprList[zbDex]+AprList[shbfDex]+AprList[xlfDex]+\
                    AprList[shfDex]+AprList[xnbtDex]+float(bzgt)-AprList[jsheDex]-AprList[kgjjDex]-AprList[kzhynjDex]-\
                        AprList[kyxDex]-AprList[kshxDex]-AprList[kylxDex]-tax
            ws2['Q3'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R3'] = round(partyFee, 2)
            # 第4行数据
            ws2['B4'] = round(MayList[gwgzDex], 2)
            ws2['C4'] = round(MayList[xjgzDex], 2)
            ws2['D4'] = round(MayList[zbDex], 2)
            ws2['E4'] = round(MayList[shbfDex], 2)
            ws2['F4'] = round(MayList[xlfDex], 2)
            ws2['G4'] = round(MayList[shfDex], 2)
            ws2['H4'] = round(MayList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I4'] = round(MayList[gtDex], 2)
            elif isFindGt is False:
                ws2['I4'] = round(float(bzgt), 2)
            ws2['J4'] = round(MayList[jsheDex], 2)
            ws2['K4'] = round(MayList[kgjjDex], 2)
            ws2['L4'] = round(MayList[kzhynjDex], 2)
            ws2['M4'] = round(MayList[kyxDex], 2)
            ws2['N4'] = round(MayList[kshxDex], 2)
            ws2['O4'] = round(MayList[kylxDex], 2)
            # ws2['S4'] = MayList[srzj3Dex]
            # ws2['T4'] = MayList[ksh2Dex]
            if isFindGt is True:
                tax = (MayList[gwgzDex]+MayList[xjgzDex]+MayList[zbDex]+MayList[shbfDex]+MayList[xlfDex]+MayList[shfDex]+\
                    MayList[xnbtDex]+MayList[gtDex])/MayList[srzj3Dex]*MayList[ksh2Dex]
            elif isFindGt is False:
                tax = (MayList[gwgzDex]+MayList[xjgzDex]+MayList[zbDex]+MayList[shbfDex]+MayList[xlfDex]+MayList[shfDex]+\
                    MayList[xnbtDex]+float(bzgt))/MayList[srzj3Dex]*MayList[ksh2Dex]
            ws2['P4'] = round(tax,2)
            if isFindGt is True:
                total = MayList[gwgzDex]+MayList[xjgzDex]+MayList[zbDex]+MayList[shbfDex]+MayList[xlfDex]+\
                    MayList[shfDex]+MayList[xnbtDex]+MayList[gtDex]-MayList[jsheDex]-MayList[kgjjDex]-MayList[kzhynjDex]-\
                        MayList[kyxDex]-MayList[kshxDex]-MayList[kylxDex]-tax
            elif isFindGt is False:
                total = MayList[gwgzDex]+MayList[xjgzDex]+MayList[zbDex]+MayList[shbfDex]+MayList[xlfDex]+\
                    MayList[shfDex]+MayList[xnbtDex]+float(bzgt)-MayList[jsheDex]-MayList[kgjjDex]-MayList[kzhynjDex]-\
                        MayList[kyxDex]-MayList[kshxDex]-MayList[kylxDex]-tax
            ws2['Q4'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R4'] = round(partyFee, 2)
            # 第5行数据
            ws2['B5'] = round(JunList[gwgzDex], 2)
            ws2['C5'] = round(JunList[xjgzDex], 2)
            ws2['D5'] = round(JunList[zbDex], 2)
            ws2['E5'] = round(JunList[shbfDex], 2)
            ws2['F5'] = round(JunList[xlfDex], 2)
            ws2['G5'] = round(JunList[shfDex], 2)
            ws2['H5'] = round(JunList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I5'] = round(JunList[gtDex], 2)
            elif isFindGt is False:
                ws2['I5'] = round(float(bzgt), 2)
            ws2['J5'] = round(JunList[jsheDex], 2)
            ws2['K5'] = round(JunList[kgjjDex], 2)
            ws2['L5'] = round(JunList[kzhynjDex], 2)
            ws2['M5'] = round(JunList[kyxDex], 2)
            ws2['N5'] = round(JunList[kshxDex], 2)
            ws2['O5'] = round(JunList[kylxDex], 2)
            if isFindGt is True:
                tax = (JunList[gwgzDex]+JunList[xjgzDex]+JunList[zbDex]+JunList[shbfDex]+JunList[xlfDex]+JunList[shfDex]+\
                    JunList[xnbtDex]+JunList[gtDex])/JunList[srzj3Dex]*JunList[ksh2Dex]
            elif isFindGt is False:
                tax = (JunList[gwgzDex]+JunList[xjgzDex]+JunList[zbDex]+JunList[shbfDex]+JunList[xlfDex]+JunList[shfDex]+\
                    JunList[xnbtDex]+float(bzgt))/JunList[srzj3Dex]*JunList[ksh2Dex]
            ws2['P5'] = round(tax, 2)
            if isFindGt is True:
                total = JunList[gwgzDex]+JunList[xjgzDex]+JunList[zbDex]+JunList[shbfDex]+JunList[xlfDex]+\
                    JunList[shfDex]+JunList[xnbtDex]+JunList[gtDex]-JunList[jsheDex]-JunList[kgjjDex]-JunList[kzhynjDex]-\
                        JunList[kyxDex]-JunList[kshxDex]-JunList[kylxDex]-tax
            elif isFindGt is False:
                total = JunList[gwgzDex]+JunList[xjgzDex]+JunList[zbDex]+JunList[shbfDex]+JunList[xlfDex]+\
                    JunList[shfDex]+JunList[xnbtDex]+float(bzgt)-JunList[jsheDex]-JunList[kgjjDex]-JunList[kzhynjDex]-\
                        JunList[kyxDex]-JunList[kshxDex]-JunList[kylxDex]-tax
            ws2['Q5'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R5'] = round(partyFee, 2)
        # 三季度
        elif season == 3:
            # 第3行数据
            ws2['B3'] = round(JulList[gwgzDex], 2)
            ws2['C3'] = round(JulList[xjgzDex], 2)
            ws2['D3'] = round(JulList[zbDex], 2)
            ws2['E3'] = round(JulList[shbfDex], 2)
            ws2['F3'] = round(JulList[xlfDex], 2)
            ws2['G3'] = round(JulList[shfDex], 2)
            ws2['H3'] = round(JulList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I3'] = round(JulList[gtDex], 2)
            elif isFindGt is False:
                ws2['I3'] = round(float(bzgt), 2)
            ws2['J3'] = round(JulList[jsheDex], 2)
            ws2['K3'] = round(JulList[kgjjDex], 2)
            ws2['L3'] = round(JulList[kzhynjDex], 2)
            ws2['M3'] = round(JulList[kyxDex], 2)
            ws2['N3'] = round(JulList[kshxDex], 2)
            ws2['O3'] = round(JulList[kylxDex], 2)
            if isFindGt is True:    
                tax = (JulList[gwgzDex]+JulList[xjgzDex]+JulList[zbDex]+JulList[shbfDex]+JulList[xlfDex]+JulList[shfDex]+\
                    JulList[xnbtDex]+JulList[gtDex])/JulList[srzj3Dex]*JulList[ksh2Dex]
            elif isFindGt is False:
                tax = (JulList[gwgzDex]+JulList[xjgzDex]+JulList[zbDex]+JulList[shbfDex]+JulList[xlfDex]+JulList[shfDex]+\
                    JulList[xnbtDex]+float(bzgt))/JulList[srzj3Dex]*JulList[ksh2Dex]
            ws2['P3'] = round(tax, 2)
            if isFindGt is True:
                total = JulList[gwgzDex]+JulList[xjgzDex]+JulList[zbDex]+JulList[shbfDex]+JulList[xlfDex]+\
                    JulList[shfDex]+JulList[xnbtDex]+JulList[gtDex]-JulList[jsheDex]-JulList[kgjjDex]-JulList[kzhynjDex]-\
                        JulList[kyxDex]-JulList[kshxDex]-JulList[kylxDex]-tax
            elif isFindGt is False:
                total = JulList[gwgzDex]+JulList[xjgzDex]+JulList[zbDex]+JulList[shbfDex]+JulList[xlfDex]+\
                    JulList[shfDex]+JulList[xnbtDex]+float(bzgt)-JulList[jsheDex]-JulList[kgjjDex]-JulList[kzhynjDex]-\
                        JulList[kyxDex]-JulList[kshxDex]-JulList[kylxDex]-tax
            ws2['Q3'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R3'] = round(partyFee, 2)
            # 第4行数据
            ws2['B4'] = round(AugList[gwgzDex], 2)
            ws2['C4'] = round(AugList[xjgzDex], 2)
            ws2['D4'] = round(AugList[zbDex], 2)
            ws2['E4'] = round(AugList[shbfDex], 2)
            ws2['F4'] = round(AugList[xlfDex], 2)
            ws2['G4'] = round(AugList[shfDex], 2)
            ws2['H4'] = round(AugList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I4'] = round(AugList[gtDex], 2)
            elif isFindGt is False:
                ws2['I4'] = round(float(bzgt), 2)
            ws2['J4'] = round(AugList[jsheDex], 2)
            ws2['K4'] = round(AugList[kgjjDex], 2)
            ws2['L4'] = round(AugList[kzhynjDex], 2)
            ws2['M4'] = round(AugList[kyxDex], 2)
            ws2['N4'] = round(AugList[kshxDex], 2)
            ws2['O4'] = round(AugList[kylxDex], 2)
            if isFindGt is True:
                tax = (AugList[gwgzDex]+AugList[xjgzDex]+AugList[zbDex]+AugList[shbfDex]+AugList[xlfDex]+AugList[shfDex]+\
                    AugList[xnbtDex]+AugList[gtDex])/AugList[srzj3Dex]*AugList[ksh2Dex]
            elif isFindGt is False:
                tax = (AugList[gwgzDex]+AugList[xjgzDex]+AugList[zbDex]+AugList[shbfDex]+AugList[xlfDex]+AugList[shfDex]+\
                    AugList[xnbtDex]+float(bzgt))/AugList[srzj3Dex]*AugList[ksh2Dex]
            ws2['P4'] = tax
            if isFindGt is True:
                total = AugList[gwgzDex]+AugList[xjgzDex]+AugList[zbDex]+AugList[shbfDex]+AugList[xlfDex]+\
                    AugList[shfDex]+AugList[xnbtDex]+AugList[gtDex]-AugList[jsheDex]-AugList[kgjjDex]-AugList[kzhynjDex]-\
                        AugList[kyxDex]-AugList[kshxDex]-AugList[kylxDex]-tax
            elif isFindGt is False:
                total = AugList[gwgzDex]+AugList[xjgzDex]+AugList[zbDex]+AugList[shbfDex]+AugList[xlfDex]+\
                    AugList[shfDex]+AugList[xnbtDex]+float(bzgt)-AugList[jsheDex]-AugList[kgjjDex]-AugList[kzhynjDex]-\
                        AugList[kyxDex]-AugList[kshxDex]-AugList[kylxDex]-tax
            ws2['Q4'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R4'] = round(partyFee, 2)
            # 第5行数据
            ws2['B5'] = round(SepList[gwgzDex], 2)
            ws2['C5'] = round(SepList[xjgzDex], 2)
            ws2['D5'] = round(SepList[zbDex], 2)
            ws2['E5'] = round(SepList[shbfDex], 2)
            ws2['F5'] = round(SepList[xlfDex], 2)
            ws2['G5'] = round(SepList[shfDex], 2)
            ws2['H5'] = round(SepList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I5'] = round(SepList[gtDex], 2)
            elif isFindGt is False:
                ws2['I5'] = round(float(bzgt), 2)
            ws2['J5'] = round(SepList[jsheDex], 2)
            ws2['K5'] = round(SepList[kgjjDex], 2)
            ws2['L5'] = round(SepList[kzhynjDex], 2)
            ws2['M5'] = round(SepList[kyxDex], 2)
            ws2['N5'] = round(SepList[kshxDex], 2)
            ws2['O5'] = round(SepList[kylxDex], 2)
            if isFindGt is True:
                tax = (SepList[gwgzDex]+SepList[xjgzDex]+SepList[zbDex]+SepList[shbfDex]+SepList[xlfDex]+SepList[shfDex]+\
                    SepList[xnbtDex]+SepList[gtDex])/SepList[srzj3Dex]*SepList[ksh2Dex]
            elif isFindGt is False:
                tax = (SepList[gwgzDex]+SepList[xjgzDex]+SepList[zbDex]+SepList[shbfDex]+SepList[xlfDex]+SepList[shfDex]+\
                    SepList[xnbtDex]+float(bzgt))/SepList[srzj3Dex]*SepList[ksh2Dex]
            ws2['P5'] = round(tax, 2)
            if isFindGt is True:
                total = SepList[gwgzDex]+SepList[xjgzDex]+SepList[zbDex]+SepList[shbfDex]+SepList[xlfDex]+\
                    SepList[shfDex]+SepList[xnbtDex]+SepList[gtDex]-SepList[jsheDex]-SepList[kgjjDex]-SepList[kzhynjDex]-\
                        SepList[kyxDex]-SepList[kshxDex]-SepList[kylxDex]-tax
            elif isFindGt is False:
                total = SepList[gwgzDex]+SepList[xjgzDex]+SepList[zbDex]+SepList[shbfDex]+SepList[xlfDex]+\
                    SepList[shfDex]+SepList[xnbtDex]+float(bzgt)-SepList[jsheDex]-SepList[kgjjDex]-SepList[kzhynjDex]-\
                        SepList[kyxDex]-SepList[kshxDex]-SepList[kylxDex]-tax
            ws2['Q5'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R5'] = round(partyFee, 2)
        # 四季度
        elif season == 4:
            # 第3行数据
            ws2['B3'] = round(OctList[gwgzDex], 2)
            ws2['C3'] = round(OctList[xjgzDex], 2)
            ws2['D3'] = round(OctList[zbDex], 2)
            ws2['E3'] = round(OctList[shbfDex], 2)
            ws2['F3'] = round(OctList[xlfDex], 2)
            ws2['G3'] = round(OctList[shfDex], 2)
            ws2['H3'] = round(OctList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I3'] = round(OctList[gtDex], 2)
            elif isFindGt is False:
                ws2['I3'] = round(float(bzgt), 2)
            ws2['J3'] = round(OctList[jsheDex], 2)
            ws2['K3'] = round(OctList[kgjjDex], 2)
            ws2['L3'] = round(OctList[kzhynjDex], 2)
            ws2['M3'] = round(OctList[kyxDex], 2)
            ws2['N3'] = round(OctList[kshxDex], 2)
            ws2['O3'] = round(OctList[kylxDex], 2)
            if isFindGt is True:
                tax = (OctList[gwgzDex]+OctList[xjgzDex]+OctList[zbDex]+OctList[shbfDex]+OctList[xlfDex]+OctList[shfDex]+\
                    OctList[xnbtDex]+OctList[gtDex])/OctList[srzj3Dex]*OctList[ksh2Dex]
            elif isFindGt is False:
                tax = (OctList[gwgzDex]+OctList[xjgzDex]+OctList[zbDex]+OctList[shbfDex]+OctList[xlfDex]+OctList[shfDex]+\
                    OctList[xnbtDex]+float(bzgt))/OctList[srzj3Dex]*OctList[ksh2Dex]
            ws2['P3'] = round(tax, 2)
            if isFindGt is True:
                total = OctList[gwgzDex]+OctList[xjgzDex]+OctList[zbDex]+OctList[shbfDex]+OctList[xlfDex]+\
                    OctList[shfDex]+OctList[xnbtDex]+OctList[gtDex]-OctList[jsheDex]-OctList[kgjjDex]-OctList[kzhynjDex]-\
                        OctList[kyxDex]-OctList[kshxDex]-OctList[kylxDex]-tax
            elif isFindGt is False:
                total = OctList[gwgzDex]+OctList[xjgzDex]+OctList[zbDex]+OctList[shbfDex]+OctList[xlfDex]+\
                    OctList[shfDex]+OctList[xnbtDex]+float(bzgt)-OctList[jsheDex]-OctList[kgjjDex]-OctList[kzhynjDex]-\
                        OctList[kyxDex]-OctList[kshxDex]-OctList[kylxDex]-tax
            ws2['Q3'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R3'] = round(partyFee, 2)
            # 第4行数据
            ws2['B4'] = round(NovList[gwgzDex], 2)
            ws2['C4'] = round(NovList[xjgzDex], 2)
            ws2['D4'] = round(NovList[zbDex], 2)
            ws2['E4'] = round(NovList[shbfDex], 2)
            ws2['F4'] = round(NovList[xlfDex], 2)
            ws2['G4'] = round(NovList[shfDex], 2)
            ws2['H4'] = round(NovList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I4'] = round(NovList[gtDex], 2)
            elif isFindGt is False:
                ws2['I4'] = round(float(bzgt), 2)
            ws2['J4'] = round(NovList[jsheDex], 2)
            ws2['K4'] = round(NovList[kgjjDex], 2)
            ws2['L4'] = round(NovList[kzhynjDex], 2)
            ws2['M4'] = round(NovList[kyxDex], 2)
            ws2['N4'] = round(NovList[kshxDex], 2)
            ws2['O4'] = round(NovList[kylxDex], 2)
            if isFindGt is True:
                tax = (NovList[gwgzDex]+NovList[xjgzDex]+NovList[zbDex]+NovList[shbfDex]+NovList[xlfDex]+NovList[shfDex]+\
                    NovList[xnbtDex]+NovList[gtDex])/NovList[srzj3Dex]*NovList[ksh2Dex]
            elif isFindGt is False:
                tax = (NovList[gwgzDex]+NovList[xjgzDex]+NovList[zbDex]+NovList[shbfDex]+NovList[xlfDex]+NovList[shfDex]+\
                    NovList[xnbtDex]+float(bzgt))/NovList[srzj3Dex]*NovList[ksh2Dex]
            ws2['P4'] = round(tax, 2)
            if isFindGt is True:
                total = NovList[gwgzDex]+NovList[xjgzDex]+NovList[zbDex]+NovList[shbfDex]+NovList[xlfDex]+\
                    NovList[shfDex]+NovList[xnbtDex]+NovList[gtDex]-NovList[jsheDex]-NovList[kgjjDex]-NovList[kzhynjDex]-\
                        NovList[kyxDex]-NovList[kshxDex]-NovList[kylxDex]-tax
            elif isFindGt is False:
                total = NovList[gwgzDex]+NovList[xjgzDex]+NovList[zbDex]+NovList[shbfDex]+NovList[xlfDex]+\
                    NovList[shfDex]+NovList[xnbtDex]+float(bzgt)-NovList[jsheDex]-NovList[kgjjDex]-NovList[kzhynjDex]-\
                        NovList[kyxDex]-NovList[kshxDex]-NovList[kylxDex]-tax
            ws2['Q4'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R4'] = round(partyFee, 2)
            # 第5行数据
            ws2['B5'] = round(DecList[gwgzDex], 2)
            ws2['C5'] = round(DecList[xjgzDex], 2)
            ws2['D5'] = round(DecList[zbDex], 2)
            ws2['E5'] = round(DecList[shbfDex], 2)
            ws2['F5'] = round(DecList[xlfDex], 2)
            ws2['G5'] = round(DecList[shfDex], 2)
            ws2['H5'] = round(DecList[xnbtDex], 2)
            if isFindGt is True:
                ws2['I5'] = round(DecList[gtDex], 2)
            elif isFindGt is False:
                ws2['I5'] = round(float(bzgt), 2)
            ws2['J5'] = round(DecList[jsheDex], 2)
            ws2['K5'] = round(DecList[kgjjDex], 2)
            ws2['L5'] = round(DecList[kzhynjDex], 2)
            ws2['M5'] = round(DecList[kyxDex], 2)
            ws2['N5'] = round(DecList[kshxDex], 2)
            ws2['O5'] = round(DecList[kylxDex], 2)
            if isFindGt is True:
                tax = (DecList[gwgzDex]+DecList[xjgzDex]+DecList[zbDex]+DecList[shbfDex]+DecList[xlfDex]+DecList[shfDex]+\
                    DecList[xnbtDex]+DecList[gtDex])/DecList[srzj3Dex]*DecList[ksh2Dex]
            elif isFindGt is False:
                tax = (DecList[gwgzDex]+DecList[xjgzDex]+DecList[zbDex]+DecList[shbfDex]+DecList[xlfDex]+DecList[shfDex]+\
                    DecList[xnbtDex]+float(bzgt))/DecList[srzj3Dex]*DecList[ksh2Dex]
            ws2['P5'] = round(tax, 2)
            if isFindGt is True:
                total = DecList[gwgzDex]+DecList[xjgzDex]+DecList[zbDex]+DecList[shbfDex]+DecList[xlfDex]+\
                    DecList[shfDex]+DecList[xnbtDex]+DecList[gtDex]-DecList[jsheDex]-DecList[kgjjDex]-DecList[kzhynjDex]-\
                        DecList[kyxDex]-DecList[kshxDex]-DecList[kylxDex]-tax
            elif isFindGt is False:
                total = DecList[gwgzDex]+DecList[xjgzDex]+DecList[zbDex]+DecList[shbfDex]+DecList[xlfDex]+\
                    DecList[shfDex]+DecList[xnbtDex]+float(bzgt)-DecList[jsheDex]-DecList[kgjjDex]-DecList[kzhynjDex]-\
                        DecList[kyxDex]-DecList[kshxDex]-DecList[kylxDex]-tax
            ws2['Q5'] = round(total, 2)
            if total <= 3000:
                partyFee = total*0.005
            elif total > 3000 and total <= 5000:
                partyFee = total*0.01
            elif total > 5000 and total <= 10000:
                partyFee = total*0.015
            elif total > 10000:
                partyFee = total*0.02
            ws2['R5'] = round(partyFee, 2)
        totalPartyFee = round(ws2['R3'].value + ws2['R4'].value + ws2['R5'].value, 2)
        ws2['B6'] = totalPartyFee
        ws2['E9'] = '应交总金额：'+str(totalPartyFee)+'元'
        ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
        ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
        # time.sleep(2)
        partyFeeName = str(year)+'年第'+str(season)+'季度_应缴党费.xlsx'
        wb2.save(current_dir + '/' + partyFeeName)
        # print()
        # print('正在生成应交党费数据...')
        # time.sleep(2)
        # input('工资数据写入成功！保存文件名为【'+partyFeeName+'】，按回车键退出程序！')
        QtWidgets.QMessageBox.information(self, "温馨提示", "缴费表格已生成！", QtWidgets.QMessageBox.Yes)
if __name__ == '__main__':
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    app = QtWidgets.QApplication(sys.argv)
    app.setAttribute(Qt.AA_EnableHighDpiScaling)
    window = mywindow()
    window.show()
    sys.exit(app.exec_())

