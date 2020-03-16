from PyQt5 import QtWidgets, QtCore, QtGui
from Ui_party_dues_gui_hd import Ui_MainWindow
import sys
import os
import datetime
import sys
import openpyxl
import xlrd
import time
import icon_qr
from openpyxl.styles import Font, colors, Border, Side, Alignment, PatternFill

class mywindow(QtWidgets.QMainWindow, Ui_MainWindow):
    
    def __init__(self):
        super(mywindow,self).__init__()
        self.setupUi(self)
        self.setFixedSize(self.width(), self.height())
        self.importButton.clicked.connect(self.fileSelect)
        self.actionImport.triggered.connect(self.fileSelect)
        self.calculateButton.clicked.connect(self.calculation)
        self.setWindowIcon(QtGui.QIcon(':/communist.png'))

    def fileSelect(self):
        filename, filetype = QtWidgets.QFileDialog.getOpenFileName(self,'选择工资文件','','Excel 2003文件(*.xls)')
        print(filename)
        filepathDisplay = str(filename).replace("/","\\")
        self.lineEdit.setText(filepathDisplay)
        global filepath 
        # filepath = str(filename).replace("/","\\")
        filepath = filename
        self.calculateButton.setEnabled(True)
    
    def calculation(self):
        def text2float(x):
            x = str(x)
            try:
                x = float(x.replace(",",""))
            except ValueError:
                pass
            return x

        def list2float(x):
            for i in range(len(x)):
                x[i] = text2float(x[i])
            return x
        
        # 判断是否自动读取岗贴
        if self.AutoGT.isChecked():
            useGtDex = True
        else:
            useGtDex = False
        
        if self.ManualGT.isChecked():
            useGtDex = False
            if self.ManagementPosition.isChecked():
                managementGTList = [11400,10200,8400,8300,7900,7600,7200,7000,6800,6300,6200,6100,5900,5800,5500,5400,5300]
                bzgt = managementGTList[self.ManagementList.currentRow()]
                print(bzgt)
            elif self.TeachingPosition.isChecked():
                teachingGTList = [14100,11600,9100,8400,7600,7200,6800,6300,6200,6100,5500,5400]
                bzgt = teachingGTList[self.TeachingList.currentRow()]
                print(bzgt)
            elif self.OtherSpecialtyPosition.isChecked():
                specilatyGTList = [8500,7700,6800,6600,6400,6100,5950,5800,5500,5400,5300]
                bzgt = specilatyGTList[self.OtherSpecialtyList.currentRow()]
                print(bzgt)
            elif self.workerPosition.isChecked():
                workerGTList = [6050,5800,5300,4950,4600,4200,4000]
                bzgt = workerGTList[self.workerList.currentRow()]
                print(bzgt)

        unitName, okPressed = QtWidgets.QInputDialog.getText(self, "单位名称", "请输入您的单位名称", QtWidgets.QLineEdit.Normal, "工商管理学院")
        # 获取当前年份，判定季度 
        year = datetime.datetime.now().strftime('%Y')
        # 获取当前月份，判定季度
        month = datetime.datetime.now().strftime('%m')
        if int(month) < 3:
            # print('\n还没有到3月，无法计算第一季度党费，程序将于5秒后自动退出...')
            QtWidgets.QMessageBox.information(self, "错误提示", "还没有到3月，无法计算第一季度党费，程序即将关闭！", QtWidgets.QMessageBox.Yes)
            exit()
        elif int(month) >= 3 and int(month) < 6:
            season = 1
            cellMonth1 = '1'
            cellMonth2 = '2'
            cellMonth3 = '3'
        elif int(month) >= 6 and int(month) < 9:
            season = 2
            cellMonth1 = '4'
            cellMonth2 = '5'
            cellMonth3 = '6'
        elif int(month) >= 9 and int(month) < 12:
            season = 3
            cellMonth1 = '7'
            cellMonth2 = '8'
            cellMonth3 = '9'
        else:
            season = 4
            cellMonth1 = '10'
            cellMonth2 = '11'
            cellMonth3 = '12'
        xlspath = filepath
        current_dir = os.path.dirname(os.path.abspath(xlspath))
        data = xlrd.open_workbook(xlspath)
        table = data.sheets()[0]
        dueNameList = table.col_values(0)
        try:
            JanList = list2float(table.col_values(1))
            FebList = list2float(table.col_values(2))
            MarList = list2float(table.col_values(3))
            AprList = list2float(table.col_values(4))
            MayList = list2float(table.col_values(5))
            JunList = list2float(table.col_values(6))
            JulList = list2float(table.col_values(7))
            AugList = list2float(table.col_values(8))
            SepList = list2float(table.col_values(9))
            OctList = list2float(table.col_values(10))
            NovList = list2float(table.col_values(11))
            DecList = list2float(table.col_values(12))
        except IndexError:
            pass
        # 缴费项目索引
        errorTips = "无法在您提供的工资表中找到【%s】项目，如果此项目缺失、或此项目名称在工资表内有变化，"\
                    "请您自行在工资表中增加该项目、或者将工资表中相应的项目名称改回为【%s】(外面不能加方括号，只留文字)，之后再重新导入工资表。\n\n点击OK后，程序将自动退出!"
        try:
            gwgzDex = dueNameList.index('岗位工资')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('岗位工资','岗位工资'), QtWidgets.QMessageBox.Ok)
            exit()
        try:
            xjgzDex = dueNameList.index('薪级工资')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('薪级工资','薪级工资'), QtWidgets.QMessageBox.Ok)
            exit()
        try:
            zbDex = dueNameList.index('职补')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('职补','职补'), QtWidgets.QMessageBox.Ok)
            exit()
        try:
            shbfDex = dueNameList.index('书报费')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('书报费','书报费'), QtWidgets.QMessageBox.Ok)
            exit()
        try:
            xlfDex = dueNameList.index('洗理费')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('洗理费','洗理费'), QtWidgets.QMessageBox.Ok)
            exit()
        try:
            shfDex = dueNameList.index('生活费')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('生活费','生活费'), QtWidgets.QMessageBox.Ok)
            exit()
        try:
            xnbtDex = dueNameList.index('校内补贴')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('校内补贴','校内补贴'), QtWidgets.QMessageBox.Ok)
            exit()
        if useGtDex is True:
            try:
                gtDex = dueNameList.index('岗位津贴')
                isFindGt = True
            except ValueError:
                try:
                    gtDex = dueNameList.index('岗贴')
                    isFindGt = True
                except ValueError:
                    isFindGt = False
                    # print('\n无法在您提供的工资表中找到【岗贴】，无法自动生成缴费表')
                    QtWidgets.QMessageBox.information(self, "错误提示", "无法在您提供的工资表中找到【岗贴】或者【岗位津贴】项目，请重启程序后，在列表中手动选择岗贴标准。"\
                                                            "\n\n点击OK后，程序将自动退出!", QtWidgets.QMessageBox.Ok)
                    exit()
                    # bzgt, okPressed = QtWidgets.QInputDialog.getInt(self, "自定义岗贴", "请输入岗贴(整数):", 0, 5400, 14100, 100)
        elif useGtDex is False:
            isFindGt = False
        try:
            jsheDex = dueNameList.index('绩效减少额')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('绩效减少额','绩效减少额'), QtWidgets.QMessageBox.Ok)
            exit()
        try:
            kgjjDex = dueNameList.index('扣公积金')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('扣公积金','扣公积金'), QtWidgets.QMessageBox.Ok)
            exit()
        try:
            kzhynjDex = dueNameList.index('扣职业年金')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('扣职业年金','扣职业年金'), QtWidgets.QMessageBox.Ok)
            exit()
        try:
            kyxDex = dueNameList.index('扣养险(新老合并）')
        except ValueError:
            try:
                kyxDex = dueNameList.index('扣养险')
            except ValueError:
                try:
                    kyxDex = dueNameList.index('扣养险(新老合并)')
                except ValueError:
                    QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('扣养险','扣养险'), QtWidgets.QMessageBox.Ok)
                    exit()
        try:    
            kshxDex = dueNameList.index('扣失险')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('扣失险','扣失险'), QtWidgets.QMessageBox.Ok)
            exit()
        try:
            kylxDex = dueNameList.index('扣医疗险')
        except ValueError:
            QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('扣医疗险','扣医疗险'), QtWidgets.QMessageBox.Ok)
            exit()
        try:
            srzj3Dex = dueNameList.index('薪金收入合计(3项)')
        except ValueError:
            try:
                srzj3Dex = dueNameList.index('薪金收入合计')
            except ValueError:
                try:
                    srzj3Dex = dueNameList.index('收入总计')
                except ValueError:
                    try:
                        srzj3Dex = dueNameList.index('收入总计(三项)')
                    except ValueError:
                        QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('薪金收入合计','薪金收入合计'), QtWidgets.QMessageBox.Ok)
                        exit()
        ksh2Text =  "检测到某些月份【%s】的数字等于0，请咨询支部书记，在计算党费时，对于【%s】项目为0的月份，是否需要"\
                    "把【扣税1】项目中的数字复制到【%s】中进行替代并参与计算。如果需要，请关闭程序，自行修改Excel原始工资表后，再重新启动程序进行计算。\n\n"\
                    "如不需要，请点击OK，本程序将把【%s】项目的数值视为0，并继续进行党费计算。计算完成后，请务必仔细核对缴费表中的数字是否正确！"
        try:
            ksh2Dex = dueNameList.index('扣税2') 
            if JanList[ksh2Dex] or FebList[ksh2Dex] or MarList[ksh2Dex] or AprList[ksh2Dex] or MayList[ksh2Dex] or JunList[ksh2Dex] or JulList[ksh2Dex] or AugList[ksh2Dex] or\
                SepList[ksh2Dex] or OctList[ksh2Dex] or NovList[ksh2Dex] or DecList[ksh2Dex] == 0:
                QtWidgets.QMessageBox.information(self, "提示", ksh2Text % ('扣税2','扣税2','扣税2','扣税2'), QtWidgets.QMessageBox.Ok)
        except ValueError:
            try:
                ksh2Dex = dueNameList.index('扣税2(本期应缴税）')
                if JanList[ksh2Dex] or FebList[ksh2Dex] or MarList[ksh2Dex] or AprList[ksh2Dex] or MayList[ksh2Dex] or JunList[ksh2Dex] or JulList[ksh2Dex] or AugList[ksh2Dex] or\
                    SepList[ksh2Dex] or OctList[ksh2Dex] or NovList[ksh2Dex] or DecList[ksh2Dex] == 0:
                    QtWidgets.QMessageBox.information(self, "提示", ksh2Text % ('扣税2(本期应缴税）','扣税2(本期应缴税）','扣税2(本期应缴税）','扣税2(本期应缴税）'), QtWidgets.QMessageBox.Ok)
            except ValueError:
                QtWidgets.QMessageBox.information(self, "错误提示", errorTips % ('扣税2','扣税2'), QtWidgets.QMessageBox.Ok)
                exit()
        # 缴费模板生成
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        for j in range(1,10):
            ws2.row_dimensions[j].height = 30    # 1-9行 行高30
        ws2.column_dimensions['A'].width = 7.28
        # 列宽B-P
        ws2.column_dimensions['B'].width = 7.02
        ws2.column_dimensions['C'].width = 7.02
        ws2.column_dimensions['D'].width = 7.02
        ws2.column_dimensions['E'].width = 7.02
        ws2.column_dimensions['F'].width = 7.02
        ws2.column_dimensions['G'].width = 7.02
        ws2.column_dimensions['H'].width = 7.02
        ws2.column_dimensions['I'].width = 7.02
        ws2.column_dimensions['J'].width = 7.02
        ws2.column_dimensions['K'].width = 7.02
        ws2.column_dimensions['L'].width = 7.02
        ws2.column_dimensions['M'].width = 7.02
        ws2.column_dimensions['N'].width = 7.02
        ws2.column_dimensions['O'].width = 7.02
        ws2.column_dimensions['P'].width = 7.02
        # 列宽Q，R
        ws2.column_dimensions['Q'].width = 9.17
        ws2.column_dimensions['R'].width = 9.17
        # 合并A1-R1,作为标题
        ws2.merge_cells('A1:R1')
        # titleCell = ws2.cell(row = 1, column = 1)
        titleCell = ws2['A1']
        titleCell.alignment = Alignment(horizontal= 'center', vertical = 'center')
        titleCell.font = Font(size=16, bold= True)
        ws2['A1'] = unitName + '教师党员应交党费统计表（' + str(year) + '年第'+str(season)+'季度）'
        # A2 年 月
        ws2['A2'] = '年 月'
        for n in range(1,19):
            # 遍历第2行的标题格式
            subTitleCell = ws2.cell(column = n, row = 2)
            subTitleCell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            subTitleCell.fill = PatternFill(fill_type='solid',start_color='FF8FBC8F',end_color='FF8FBC8F')
            subTitleCell.border = Border(left=Side(border_style='thin',color='FF000000'),
            right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='thin',color='FF000000'),
            bottom=Side(border_style='thin',color='FF000000')
            )
            # 遍历第9行的格式
            bottomCell = ws2.cell(column = n, row = 9)
            bottomCell.alignment = Alignment(horizontal='left', vertical='center')
            bottomCell.border = Border(left=Side(border_style='thin',color='FF000000'),
            right=Side(border_style='thin',color='FF000000'),
            top=Side(border_style='thin',color='FF000000'),
            bottom=Side(border_style='thin',color='FF000000')
            )
            for m in range(3,9):
                # 遍历第3、4、5行的数据格式
                contentCell = ws2.cell(column = n, row = m)
                contentCell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                contentCell.border = Border(left=Side(border_style='thin',color='FF000000'),
                right=Side(border_style='thin',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='thin',color='FF000000')
                )
        ws2['A3'] = str(year) + '年' + cellMonth1 + '月'
        ws2['A4'] = str(year) + '年' + cellMonth2 + '月'
        ws2['A5'] = str(year) + '年' + cellMonth3 + '月'
        ws2['A6'] = '总计'
        # 合并B6-R1,作为总计数据
        ws2.merge_cells('B6:R6')
        # 合并A7-R7，作为公式展示
        ws2.merge_cells('A7:R7')
        ws2['A7'] = '所得税计算公式：所得税=（岗位工资+薪级工资+职务补贴+书报费+洗理费+生活费+校内补贴+岗位津贴）/薪金收入合计(3项)*扣税2'
        # 合并A8:B8，作为党费收缴标准
        ws2.merge_cells('A8:B8')
        # 合并C8：R8，作为标准展示
        ws2.merge_cells('C8:R8')
        ws2['A8'] = '党费收交标准：'
        ws2['C8'] = '3000元（含）以下：0.5%，3000-5000元（含）：1%，5000元-10000元（含）：1.5%，10000元以上：2%'
        # 合并A9：D9，作为姓名（签字）
        ws2.merge_cells('A9:D9')
        ws2['A9'] = '姓名（签字）：'
        # 合并E9：H9，作为应交总金额
        ws2.merge_cells('E9:H9')
        # 合并I9：M9，作为所在支部
        ws2.merge_cells('I9:M9')
        ws2['I9'] = '所在支部：'
        # 合并N9：R9，作为支部书记（签字）
        ws2.merge_cells('N9:R9')
        ws2['N9'] = '支部书记（签字）：'
        ws2['B2'] = '岗位工资'
        ws2['C2'] = '+薪级工资'
        ws2['D2'] = '+职务补贴'
        ws2['E2'] = '+书报费'
        ws2['F2'] = '+洗理费'
        ws2['G2'] = '+生活费'
        ws2['H2'] = '+校内补贴'
        ws2['I2'] = '+岗位津贴'
        ws2['J2'] = '-减少额'
        ws2['K2'] = '-扣公积金'
        ws2['L2'] = '-扣职业年金'
        ws2['M2'] = '-扣养险'
        ws2['N2'] = '-扣失险'
        ws2['O2'] = '-扣医疗险'
        ws2['P2'] = '-所得税'
        ws2['Q2'] = '合计'
        ws2['R2'] = '应交党费'
        try:
            # 一季度
            if season == 1:
                # 第3行数据
                ws2['B3'] = round(JanList[gwgzDex], 2)
                ws2['C3'] = round(JanList[xjgzDex], 2)
                ws2['D3'] = round(JanList[zbDex], 2)
                ws2['E3'] = round(JanList[shbfDex], 2)
                ws2['F3'] = round(JanList[xlfDex], 2)
                ws2['G3'] = round(JanList[shfDex], 2)
                ws2['H3'] = round(JanList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I3'] = round(text2float(JanList[gtDex]), 2)
                elif isFindGt is False:
                    ws2['I3'] = round(float(bzgt), 2)
                ws2['J3'] = round(abs(JanList[jsheDex]), 2)
                ws2['K3'] = round(abs(JanList[kgjjDex]), 2)
                ws2['L3'] = round(abs(JanList[kzhynjDex]), 2)
                ws2['M3'] = round(abs(JanList[kyxDex]), 2)
                ws2['N3'] = round(abs(JanList[kshxDex]), 2)
                ws2['O3'] = round(abs(JanList[kylxDex]), 2)
                if isFindGt is True:
                    tax = (JanList[gwgzDex]+JanList[xjgzDex]+JanList[zbDex]+JanList[shbfDex]+JanList[xlfDex]+JanList[shfDex]+\
                        JanList[xnbtDex]+JanList[gtDex])/JanList[srzj3Dex]*JanList[ksh2Dex]
                elif isFindGt is False:
                    tax = (JanList[gwgzDex]+JanList[xjgzDex]+JanList[zbDex]+JanList[shbfDex]+JanList[xlfDex]+JanList[shfDex]+\
                        JanList[xnbtDex]+float(bzgt))/JanList[srzj3Dex]*JanList[ksh2Dex]        
                ws2['P3'] = round(tax, 2)
                if isFindGt is True:
                    total = JanList[gwgzDex]+JanList[xjgzDex]+JanList[zbDex]+JanList[shbfDex]+JanList[xlfDex]+\
                        JanList[shfDex]+JanList[xnbtDex]+JanList[gtDex]-abs(JanList[jsheDex])-abs(JanList[kgjjDex])-abs(JanList[kzhynjDex])-\
                            abs(JanList[kyxDex])-abs(JanList[kshxDex])-abs(JanList[kylxDex])-tax
                elif isFindGt is False:
                    total = JanList[gwgzDex]+JanList[xjgzDex]+JanList[zbDex]+JanList[shbfDex]+JanList[xlfDex]+\
                        JanList[shfDex]+JanList[xnbtDex]+float(bzgt)-abs(JanList[jsheDex])-abs(JanList[kgjjDex])-abs(JanList[kzhynjDex])-\
                            abs(JanList[kyxDex])-abs(JanList[kshxDex])-abs(JanList[kylxDex])-tax
                ws2['Q3'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R3'] = round(partyFee, 2)
                # ws2['S3'] = JanList[srzj3Dex]
                # ws2['T3'] = JanList[ksh2Dex]
                # 第4行数据
                ws2['B4'] = round(FebList[gwgzDex], 2)
                ws2['C4'] = round(FebList[xjgzDex], 2)
                ws2['D4'] = round(FebList[zbDex], 2)
                ws2['E4'] = round(FebList[shbfDex], 2)
                ws2['F4'] = round(FebList[xlfDex], 2)
                ws2['G4'] = round(FebList[shfDex], 2)
                ws2['H4'] = round(FebList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I4'] = round(FebList[gtDex], 2)
                elif isFindGt is False:
                    ws2['I4'] = round(float(bzgt), 2)
                ws2['J4'] = round(abs(FebList[jsheDex]), 2)
                ws2['K4'] = round(abs(FebList[kgjjDex]), 2)
                ws2['L4'] = round(abs(FebList[kzhynjDex]), 2)
                ws2['M4'] = round(abs(FebList[kyxDex]), 2)
                ws2['N4'] = round(abs(FebList[kshxDex]), 2)
                ws2['O4'] = round(abs(FebList[kylxDex]), 2)
                if isFindGt is True:
                    tax = (FebList[gwgzDex]+FebList[xjgzDex]+FebList[zbDex]+FebList[shbfDex]+FebList[xlfDex]+FebList[shfDex]+\
                        FebList[xnbtDex]+FebList[gtDex])/FebList[srzj3Dex]*FebList[ksh2Dex]
                elif isFindGt is False:
                    tax = (FebList[gwgzDex]+FebList[xjgzDex]+FebList[zbDex]+FebList[shbfDex]+FebList[xlfDex]+FebList[shfDex]+\
                        FebList[xnbtDex]+float(bzgt))/FebList[srzj3Dex]*FebList[ksh2Dex]
                ws2['P4'] = round(tax, 2)
                if isFindGt is True:
                    total = FebList[gwgzDex]+FebList[xjgzDex]+FebList[zbDex]+FebList[shbfDex]+FebList[xlfDex]+\
                        FebList[shfDex]+FebList[xnbtDex]+FebList[gtDex]-abs(FebList[jsheDex])-abs(FebList[kgjjDex])-abs(FebList[kzhynjDex])-\
                            abs(FebList[kyxDex])-abs(FebList[kshxDex])-abs(FebList[kylxDex])-tax
                elif isFindGt is False:
                    total = FebList[gwgzDex]+FebList[xjgzDex]+FebList[zbDex]+FebList[shbfDex]+FebList[xlfDex]+\
                    FebList[shfDex]+FebList[xnbtDex]+float(bzgt)-abs(FebList[jsheDex])-abs(FebList[kgjjDex])-abs(FebList[kzhynjDex])-\
                        abs(FebList[kyxDex])-abs(FebList[kshxDex])-abs(FebList[kylxDex])-tax
                ws2['Q4'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R4'] = round(partyFee, 2)
                # ws2['S4'] = FebList[srzj3Dex]
                # ws2['T4'] = FebList[ksh2Dex]
                # 第5行数据
                ws2['B5'] = round(MarList[gwgzDex], 2)
                ws2['C5'] = round(MarList[xjgzDex], 2)
                ws2['D5'] = round(MarList[zbDex], 2)
                ws2['E5'] = round(MarList[shbfDex], 2)
                ws2['F5'] = round(MarList[xlfDex], 2)
                ws2['G5'] = round(MarList[shfDex], 2)
                ws2['H5'] = round(MarList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I5'] = round(MarList[gtDex], 2)
                elif isFindGt is False:
                    ws2['I5'] = round(float(bzgt), 2)
                ws2['J5'] = round(abs(MarList[jsheDex]), 2)
                ws2['K5'] = round(abs(MarList[kgjjDex]), 2)
                ws2['L5'] = round(abs(MarList[kzhynjDex]), 2)
                ws2['M5'] = round(abs(MarList[kyxDex]), 2)
                ws2['N5'] = round(abs(MarList[kshxDex]), 2)
                ws2['O5'] = round(abs(MarList[kylxDex]), 2)
                if isFindGt is True:
                    tax = (MarList[gwgzDex]+MarList[xjgzDex]+MarList[zbDex]+MarList[shbfDex]+MarList[xlfDex]+MarList[shfDex]+\
                        MarList[xnbtDex]+MarList[gtDex])/MarList[srzj3Dex]*MarList[ksh2Dex]
                elif isFindGt is False:
                    tax = (MarList[gwgzDex]+MarList[xjgzDex]+MarList[zbDex]+MarList[shbfDex]+MarList[xlfDex]+MarList[shfDex]+\
                        MarList[xnbtDex]+float(bzgt))/MarList[srzj3Dex]*MarList[ksh2Dex]
                ws2['P5'] = round(tax, 2)
                if isFindGt is True:
                    total = MarList[gwgzDex]+MarList[xjgzDex]+MarList[zbDex]+MarList[shbfDex]+MarList[xlfDex]+\
                        MarList[shfDex]+MarList[xnbtDex]+MarList[gtDex]-abs(MarList[jsheDex])-abs(MarList[kgjjDex])-abs(MarList[kzhynjDex])-\
                            abs(MarList[kyxDex])-abs(MarList[kshxDex])-abs(MarList[kylxDex])-tax
                elif isFindGt is False:
                    total = MarList[gwgzDex]+MarList[xjgzDex]+MarList[zbDex]+MarList[shbfDex]+MarList[xlfDex]+\
                        MarList[shfDex]+MarList[xnbtDex]+float(bzgt)-abs(MarList[jsheDex])-abs(MarList[kgjjDex])-abs(MarList[kzhynjDex])-\
                            abs(MarList[kyxDex])-abs(MarList[kshxDex])-abs(MarList[kylxDex])-tax
                ws2['Q5'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R5'] = round(partyFee, 2)
                # ws2['S5'] = MarList[srzj3Dex]
                # ws2['T5'] = MarList[ksh2Dex]
            # 二季度
            elif season == 2:
                # 第3行数据
                ws2['B3'] = round(AprList[gwgzDex], 2)
                ws2['C3'] = round(AprList[xjgzDex], 2)
                ws2['D3'] = round(AprList[zbDex], 2)
                ws2['E3'] = round(AprList[shbfDex], 2)
                ws2['F3'] = round(AprList[xlfDex], 2)
                ws2['G3'] = round(AprList[shfDex], 2)
                ws2['H3'] = round(AprList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I3'] = round(AprList[gtDex], 2)
                elif isFindGt is False:
                    ws2['I3'] = round(float(bzgt), 2)
                ws2['J3'] = round(abs(AprList[jsheDex]), 2)
                ws2['K3'] = round(abs(AprList[kgjjDex]), 2)
                ws2['L3'] = round(abs(AprList[kzhynjDex]), 2)
                ws2['M3'] = round(abs(AprList[kyxDex]), 2)
                ws2['N3'] = round(abs(AprList[kshxDex]), 2)
                ws2['O3'] = round(abs(AprList[kylxDex]), 2)
                # ws2['S3'] = AprList[srzj3Dex]
                # ws2['T3'] = AprList[ksh2Dex]
                if isFindGt is True:
                    tax = (AprList[gwgzDex]+AprList[xjgzDex]+AprList[zbDex]+AprList[shbfDex]+AprList[xlfDex]+AprList[shfDex]+\
                        AprList[xnbtDex]+AprList[gtDex])/AprList[srzj3Dex]*AprList[ksh2Dex]
                elif isFindGt is False:
                    tax = (AprList[gwgzDex]+AprList[xjgzDex]+AprList[zbDex]+AprList[shbfDex]+AprList[xlfDex]+AprList[shfDex]+\
                        AprList[xnbtDex]+float(bzgt))/AprList[srzj3Dex]*AprList[ksh2Dex]
                ws2['P3'] = round(tax, 2)
                if isFindGt is True:
                    total = AprList[gwgzDex]+AprList[xjgzDex]+AprList[zbDex]+AprList[shbfDex]+AprList[xlfDex]+\
                        AprList[shfDex]+AprList[xnbtDex]+AprList[gtDex]-abs(AprList[jsheDex])-abs(AprList[kgjjDex])-abs(AprList[kzhynjDex])-\
                            abs(AprList[kyxDex])-abs(AprList[kshxDex])-abs(AprList[kylxDex])-tax
                elif isFindGt is False:
                    total = AprList[gwgzDex]+AprList[xjgzDex]+AprList[zbDex]+AprList[shbfDex]+AprList[xlfDex]+\
                        AprList[shfDex]+AprList[xnbtDex]+float(bzgt)-abs(AprList[jsheDex])-abs(AprList[kgjjDex])-abs(AprList[kzhynjDex])-\
                            abs(AprList[kyxDex])-abs(AprList[kshxDex])-abs(AprList[kylxDex])-tax
                ws2['Q3'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R3'] = round(partyFee, 2)
                # 第4行数据
                ws2['B4'] = round(MayList[gwgzDex], 2)
                ws2['C4'] = round(MayList[xjgzDex], 2)
                ws2['D4'] = round(MayList[zbDex], 2)
                ws2['E4'] = round(MayList[shbfDex], 2)
                ws2['F4'] = round(MayList[xlfDex], 2)
                ws2['G4'] = round(MayList[shfDex], 2)
                ws2['H4'] = round(MayList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I4'] = round(MayList[gtDex], 2)
                elif isFindGt is False:
                    ws2['I4'] = round(float(bzgt), 2)
                ws2['J4'] = round(abs(MayList[jsheDex]), 2)
                ws2['K4'] = round(abs(MayList[kgjjDex]), 2)
                ws2['L4'] = round(abs(MayList[kzhynjDex]), 2)
                ws2['M4'] = round(abs(MayList[kyxDex]), 2)
                ws2['N4'] = round(abs(MayList[kshxDex]), 2)
                ws2['O4'] = round(abs(MayList[kylxDex]), 2)
                # ws2['S4'] = MayList[srzj3Dex]
                # ws2['T4'] = MayList[ksh2Dex]
                if isFindGt is True:
                    tax = (MayList[gwgzDex]+MayList[xjgzDex]+MayList[zbDex]+MayList[shbfDex]+MayList[xlfDex]+MayList[shfDex]+\
                        MayList[xnbtDex]+MayList[gtDex])/MayList[srzj3Dex]*MayList[ksh2Dex]
                elif isFindGt is False:
                    tax = (MayList[gwgzDex]+MayList[xjgzDex]+MayList[zbDex]+MayList[shbfDex]+MayList[xlfDex]+MayList[shfDex]+\
                        MayList[xnbtDex]+float(bzgt))/MayList[srzj3Dex]*MayList[ksh2Dex]
                ws2['P4'] = round(tax,2)
                if isFindGt is True:
                    total = MayList[gwgzDex]+MayList[xjgzDex]+MayList[zbDex]+MayList[shbfDex]+MayList[xlfDex]+\
                        MayList[shfDex]+MayList[xnbtDex]+MayList[gtDex]-abs(MayList[jsheDex])-abs(MayList[kgjjDex])-abs(MayList[kzhynjDex])-\
                            abs(MayList[kyxDex])-abs(MayList[kshxDex])-abs(MayList[kylxDex])-tax
                elif isFindGt is False:
                    total = MayList[gwgzDex]+MayList[xjgzDex]+MayList[zbDex]+MayList[shbfDex]+MayList[xlfDex]+\
                        MayList[shfDex]+MayList[xnbtDex]+float(bzgt)-abs(MayList[jsheDex])-abs(MayList[kgjjDex])-abs(MayList[kzhynjDex])-\
                            abs(MayList[kyxDex])-abs(MayList[kshxDex])-abs(MayList[kylxDex])-tax
                ws2['Q4'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R4'] = round(partyFee, 2)
                # 第5行数据
                ws2['B5'] = round(JunList[gwgzDex], 2)
                ws2['C5'] = round(JunList[xjgzDex], 2)
                ws2['D5'] = round(JunList[zbDex], 2)
                ws2['E5'] = round(JunList[shbfDex], 2)
                ws2['F5'] = round(JunList[xlfDex], 2)
                ws2['G5'] = round(JunList[shfDex], 2)
                ws2['H5'] = round(JunList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I5'] = round(JunList[gtDex], 2)
                elif isFindGt is False:
                    ws2['I5'] = round(float(bzgt), 2)
                ws2['J5'] = round(abs(JunList[jsheDex]), 2)
                ws2['K5'] = round(abs(JunList[kgjjDex]), 2)
                ws2['L5'] = round(abs(JunList[kzhynjDex]), 2)
                ws2['M5'] = round(abs(JunList[kyxDex]), 2)
                ws2['N5'] = round(abs(JunList[kshxDex]), 2)
                ws2['O5'] = round(abs(JunList[kylxDex]), 2)
                if isFindGt is True:
                    tax = (JunList[gwgzDex]+JunList[xjgzDex]+JunList[zbDex]+JunList[shbfDex]+JunList[xlfDex]+JunList[shfDex]+\
                        JunList[xnbtDex]+JunList[gtDex])/JunList[srzj3Dex]*JunList[ksh2Dex]
                elif isFindGt is False:
                    tax = (JunList[gwgzDex]+JunList[xjgzDex]+JunList[zbDex]+JunList[shbfDex]+JunList[xlfDex]+JunList[shfDex]+\
                        JunList[xnbtDex]+float(bzgt))/JunList[srzj3Dex]*JunList[ksh2Dex]
                ws2['P5'] = round(tax, 2)
                if isFindGt is True:
                    total = JunList[gwgzDex]+JunList[xjgzDex]+JunList[zbDex]+JunList[shbfDex]+JunList[xlfDex]+\
                        JunList[shfDex]+JunList[xnbtDex]+JunList[gtDex]-abs(JunList[jsheDex])-abs(JunList[kgjjDex])-abs(JunList[kzhynjDex])-\
                            abs(JunList[kyxDex])-abs(JunList[kshxDex])-abs(JunList[kylxDex])-tax
                elif isFindGt is False:
                    total = JunList[gwgzDex]+JunList[xjgzDex]+JunList[zbDex]+JunList[shbfDex]+JunList[xlfDex]+\
                        JunList[shfDex]+JunList[xnbtDex]+float(bzgt)-abs(JunList[jsheDex])-abs(JunList[kgjjDex])-abs(JunList[kzhynjDex])-\
                            abs(JunList[kyxDex])-abs(JunList[kshxDex])-abs(JunList[kylxDex])-tax
                ws2['Q5'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R5'] = round(partyFee, 2)
            # 三季度
            elif season == 3:
                # 第3行数据
                ws2['B3'] = round(JulList[gwgzDex], 2)
                ws2['C3'] = round(JulList[xjgzDex], 2)
                ws2['D3'] = round(JulList[zbDex], 2)
                ws2['E3'] = round(JulList[shbfDex], 2)
                ws2['F3'] = round(JulList[xlfDex], 2)
                ws2['G3'] = round(JulList[shfDex], 2)
                ws2['H3'] = round(JulList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I3'] = round(JulList[gtDex], 2)
                elif isFindGt is False:
                    ws2['I3'] = round(float(bzgt), 2)
                ws2['J3'] = round(abs(JulList[jsheDex]), 2)
                ws2['K3'] = round(abs(JulList[kgjjDex]), 2)
                ws2['L3'] = round(abs(JulList[kzhynjDex]), 2)
                ws2['M3'] = round(abs(JulList[kyxDex]), 2)
                ws2['N3'] = round(abs(JulList[kshxDex]), 2)
                ws2['O3'] = round(abs(JulList[kylxDex]), 2)
                if isFindGt is True:    
                    tax = (JulList[gwgzDex]+JulList[xjgzDex]+JulList[zbDex]+JulList[shbfDex]+JulList[xlfDex]+JulList[shfDex]+\
                        JulList[xnbtDex]+JulList[gtDex])/JulList[srzj3Dex]*JulList[ksh2Dex]
                elif isFindGt is False:
                    tax = (JulList[gwgzDex]+JulList[xjgzDex]+JulList[zbDex]+JulList[shbfDex]+JulList[xlfDex]+JulList[shfDex]+\
                        JulList[xnbtDex]+float(bzgt))/JulList[srzj3Dex]*JulList[ksh2Dex]
                ws2['P3'] = round(tax, 2)
                if isFindGt is True:
                    total = JulList[gwgzDex]+JulList[xjgzDex]+JulList[zbDex]+JulList[shbfDex]+JulList[xlfDex]+\
                        JulList[shfDex]+JulList[xnbtDex]+JulList[gtDex]-abs(JulList[jsheDex])-abs(JulList[kgjjDex])-abs(JulList[kzhynjDex])-\
                            abs(JulList[kyxDex])-abs(JulList[kshxDex])-abs(JulList[kylxDex])-tax
                elif isFindGt is False:
                    total = JulList[gwgzDex]+JulList[xjgzDex]+JulList[zbDex]+JulList[shbfDex]+JulList[xlfDex]+\
                        JulList[shfDex]+JulList[xnbtDex]+float(bzgt)-abs(JulList[jsheDex])-abs(JulList[kgjjDex])-abs(JulList[kzhynjDex])-\
                            abs(JulList[kyxDex])-abs(JulList[kshxDex])-abs(JulList[kylxDex])-tax
                ws2['Q3'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R3'] = round(partyFee, 2)
                # 第4行数据
                ws2['B4'] = round(AugList[gwgzDex], 2)
                ws2['C4'] = round(AugList[xjgzDex], 2)
                ws2['D4'] = round(AugList[zbDex], 2)
                ws2['E4'] = round(AugList[shbfDex], 2)
                ws2['F4'] = round(AugList[xlfDex], 2)
                ws2['G4'] = round(AugList[shfDex], 2)
                ws2['H4'] = round(AugList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I4'] = round(AugList[gtDex], 2)
                elif isFindGt is False:
                    ws2['I4'] = round(float(bzgt), 2)
                ws2['J4'] = round(abs(AugList[jsheDex]), 2)
                ws2['K4'] = round(abs(AugList[kgjjDex]), 2)
                ws2['L4'] = round(abs(AugList[kzhynjDex]), 2)
                ws2['M4'] = round(abs(AugList[kyxDex]), 2)
                ws2['N4'] = round(abs(AugList[kshxDex]), 2)
                ws2['O4'] = round(abs(AugList[kylxDex]), 2)
                if isFindGt is True:
                    tax = (AugList[gwgzDex]+AugList[xjgzDex]+AugList[zbDex]+AugList[shbfDex]+AugList[xlfDex]+AugList[shfDex]+\
                        AugList[xnbtDex]+AugList[gtDex])/AugList[srzj3Dex]*AugList[ksh2Dex]
                elif isFindGt is False:
                    tax = (AugList[gwgzDex]+AugList[xjgzDex]+AugList[zbDex]+AugList[shbfDex]+AugList[xlfDex]+AugList[shfDex]+\
                        AugList[xnbtDex]+float(bzgt))/AugList[srzj3Dex]*AugList[ksh2Dex]
                ws2['P4'] = tax
                if isFindGt is True:
                    total = AugList[gwgzDex]+AugList[xjgzDex]+AugList[zbDex]+AugList[shbfDex]+AugList[xlfDex]+\
                        AugList[shfDex]+AugList[xnbtDex]+AugList[gtDex]-abs(AugList[jsheDex])-abs(AugList[kgjjDex])-abs(AugList[kzhynjDex])-\
                            abs(AugList[kyxDex])-abs(AugList[kshxDex])-abs(AugList[kylxDex])-tax
                elif isFindGt is False:
                    total = AugList[gwgzDex]+AugList[xjgzDex]+AugList[zbDex]+AugList[shbfDex]+AugList[xlfDex]+\
                        AugList[shfDex]+AugList[xnbtDex]+float(bzgt)-abs(AugList[jsheDex])-abs(AugList[kgjjDex])-abs(AugList[kzhynjDex])-\
                            abs(AugList[kyxDex])-abs(AugList[kshxDex])-abs(AugList[kylxDex])-tax
                ws2['Q4'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R4'] = round(partyFee, 2)
                # 第5行数据
                ws2['B5'] = round(SepList[gwgzDex], 2)
                ws2['C5'] = round(SepList[xjgzDex], 2)
                ws2['D5'] = round(SepList[zbDex], 2)
                ws2['E5'] = round(SepList[shbfDex], 2)
                ws2['F5'] = round(SepList[xlfDex], 2)
                ws2['G5'] = round(SepList[shfDex], 2)
                ws2['H5'] = round(SepList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I5'] = round(SepList[gtDex], 2)
                elif isFindGt is False:
                    ws2['I5'] = round(float(bzgt), 2)
                ws2['J5'] = round(abs(SepList[jsheDex]), 2)
                ws2['K5'] = round(abs(SepList[kgjjDex]), 2)
                ws2['L5'] = round(abs(SepList[kzhynjDex]), 2)
                ws2['M5'] = round(abs(SepList[kyxDex]), 2)
                ws2['N5'] = round(abs(SepList[kshxDex]), 2)
                ws2['O5'] = round(abs(SepList[kylxDex]), 2)
                if isFindGt is True:
                    tax = (SepList[gwgzDex]+SepList[xjgzDex]+SepList[zbDex]+SepList[shbfDex]+SepList[xlfDex]+SepList[shfDex]+\
                        SepList[xnbtDex]+SepList[gtDex])/SepList[srzj3Dex]*SepList[ksh2Dex]
                elif isFindGt is False:
                    tax = (SepList[gwgzDex]+SepList[xjgzDex]+SepList[zbDex]+SepList[shbfDex]+SepList[xlfDex]+SepList[shfDex]+\
                        SepList[xnbtDex]+float(bzgt))/SepList[srzj3Dex]*SepList[ksh2Dex]
                ws2['P5'] = round(tax, 2)
                if isFindGt is True:
                    total = SepList[gwgzDex]+SepList[xjgzDex]+SepList[zbDex]+SepList[shbfDex]+SepList[xlfDex]+\
                        SepList[shfDex]+SepList[xnbtDex]+SepList[gtDex]-abs(SepList[jsheDex])-abs(SepList[kgjjDex])-abs(SepList[kzhynjDex])-\
                            abs(SepList[kyxDex])-abs(SepList[kshxDex])-abs(SepList[kylxDex])-tax
                elif isFindGt is False:
                    total = SepList[gwgzDex]+SepList[xjgzDex]+SepList[zbDex]+SepList[shbfDex]+SepList[xlfDex]+\
                        SepList[shfDex]+SepList[xnbtDex]+float(bzgt)-abs(SepList[jsheDex])-abs(SepList[kgjjDex])-abs(SepList[kzhynjDex])-\
                            abs(SepList[kyxDex])-abs(SepList[kshxDex])-abs(SepList[kylxDex])-tax
                ws2['Q5'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R5'] = round(partyFee, 2)
            # 四季度
            elif season == 4:
                # 第3行数据
                ws2['B3'] = round(OctList[gwgzDex], 2)
                ws2['C3'] = round(OctList[xjgzDex], 2)
                ws2['D3'] = round(OctList[zbDex], 2)
                ws2['E3'] = round(OctList[shbfDex], 2)
                ws2['F3'] = round(OctList[xlfDex], 2)
                ws2['G3'] = round(OctList[shfDex], 2)
                ws2['H3'] = round(OctList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I3'] = round(OctList[gtDex], 2)
                elif isFindGt is False:
                    ws2['I3'] = round(float(bzgt), 2)
                ws2['J3'] = round(abs(OctList[jsheDex]), 2)
                ws2['K3'] = round(abs(OctList[kgjjDex]), 2)
                ws2['L3'] = round(abs(OctList[kzhynjDex]), 2)
                ws2['M3'] = round(abs(OctList[kyxDex]), 2)
                ws2['N3'] = round(abs(OctList[kshxDex]), 2)
                ws2['O3'] = round(abs(OctList[kylxDex]), 2)
                if isFindGt is True:
                    tax = (OctList[gwgzDex]+OctList[xjgzDex]+OctList[zbDex]+OctList[shbfDex]+OctList[xlfDex]+OctList[shfDex]+\
                        OctList[xnbtDex]+OctList[gtDex])/OctList[srzj3Dex]*OctList[ksh2Dex]
                elif isFindGt is False:
                    tax = (OctList[gwgzDex]+OctList[xjgzDex]+OctList[zbDex]+OctList[shbfDex]+OctList[xlfDex]+OctList[shfDex]+\
                        OctList[xnbtDex]+float(bzgt))/OctList[srzj3Dex]*OctList[ksh2Dex]
                ws2['P3'] = round(tax, 2)
                if isFindGt is True:
                    total = OctList[gwgzDex]+OctList[xjgzDex]+OctList[zbDex]+OctList[shbfDex]+OctList[xlfDex]+\
                        OctList[shfDex]+OctList[xnbtDex]+OctList[gtDex]-abs(OctList[jsheDex])-abs(OctList[kgjjDex])-abs(OctList[kzhynjDex])-\
                            abs(OctList[kyxDex])-abs(OctList[kshxDex])-abs(OctList[kylxDex])-tax
                elif isFindGt is False:
                    total = OctList[gwgzDex]+OctList[xjgzDex]+OctList[zbDex]+OctList[shbfDex]+OctList[xlfDex]+\
                        OctList[shfDex]+OctList[xnbtDex]+float(bzgt)-abs(OctList[jsheDex])-abs(OctList[kgjjDex])-abs(OctList[kzhynjDex])-\
                            abs(OctList[kyxDex])-abs(OctList[kshxDex])-abs(OctList[kylxDex])-tax
                ws2['Q3'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R3'] = round(partyFee, 2)
                # 第4行数据
                ws2['B4'] = round(NovList[gwgzDex], 2)
                ws2['C4'] = round(NovList[xjgzDex], 2)
                ws2['D4'] = round(NovList[zbDex], 2)
                ws2['E4'] = round(NovList[shbfDex], 2)
                ws2['F4'] = round(NovList[xlfDex], 2)
                ws2['G4'] = round(NovList[shfDex], 2)
                ws2['H4'] = round(NovList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I4'] = round(NovList[gtDex], 2)
                elif isFindGt is False:
                    ws2['I4'] = round(float(bzgt), 2)
                ws2['J4'] = round(abs(NovList[jsheDex]), 2)
                ws2['K4'] = round(abs(NovList[kgjjDex]), 2)
                ws2['L4'] = round(abs(NovList[kzhynjDex]), 2)
                ws2['M4'] = round(abs(NovList[kyxDex]), 2)
                ws2['N4'] = round(abs(NovList[kshxDex]), 2)
                ws2['O4'] = round(abs(NovList[kylxDex]), 2)
                if isFindGt is True:
                    tax = (NovList[gwgzDex]+NovList[xjgzDex]+NovList[zbDex]+NovList[shbfDex]+NovList[xlfDex]+NovList[shfDex]+\
                        NovList[xnbtDex]+NovList[gtDex])/NovList[srzj3Dex]*NovList[ksh2Dex]
                elif isFindGt is False:
                    tax = (NovList[gwgzDex]+NovList[xjgzDex]+NovList[zbDex]+NovList[shbfDex]+NovList[xlfDex]+NovList[shfDex]+\
                        NovList[xnbtDex]+float(bzgt))/NovList[srzj3Dex]*NovList[ksh2Dex]
                ws2['P4'] = round(tax, 2)
                if isFindGt is True:
                    total = NovList[gwgzDex]+NovList[xjgzDex]+NovList[zbDex]+NovList[shbfDex]+NovList[xlfDex]+\
                        NovList[shfDex]+NovList[xnbtDex]+NovList[gtDex]-abs(NovList[jsheDex])-abs(NovList[kgjjDex])-abs(NovList[kzhynjDex])-\
                            abs(NovList[kyxDex])-abs(NovList[kshxDex])-abs(NovList[kylxDex])-tax
                elif isFindGt is False:
                    total = NovList[gwgzDex]+NovList[xjgzDex]+NovList[zbDex]+NovList[shbfDex]+NovList[xlfDex]+\
                        NovList[shfDex]+NovList[xnbtDex]+float(bzgt)-abs(NovList[jsheDex])-abs(NovList[kgjjDex])-abs(NovList[kzhynjDex])-\
                            abs(NovList[kyxDex])-abs(NovList[kshxDex])-abs(NovList[kylxDex])-tax
                ws2['Q4'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R4'] = round(partyFee, 2)
                # 第5行数据
                ws2['B5'] = round(DecList[gwgzDex], 2)
                ws2['C5'] = round(DecList[xjgzDex], 2)
                ws2['D5'] = round(DecList[zbDex], 2)
                ws2['E5'] = round(DecList[shbfDex], 2)
                ws2['F5'] = round(DecList[xlfDex], 2)
                ws2['G5'] = round(DecList[shfDex], 2)
                ws2['H5'] = round(DecList[xnbtDex], 2)
                if isFindGt is True:
                    ws2['I5'] = round(DecList[gtDex], 2)
                elif isFindGt is False:
                    ws2['I5'] = round(float(bzgt), 2)
                ws2['J5'] = round(abs(DecList[jsheDex]), 2)
                ws2['K5'] = round(abs(DecList[kgjjDex]), 2)
                ws2['L5'] = round(abs(DecList[kzhynjDex]), 2)
                ws2['M5'] = round(abs(DecList[kyxDex]), 2)
                ws2['N5'] = round(abs(DecList[kshxDex]), 2)
                ws2['O5'] = round(abs(DecList[kylxDex]), 2)
                if isFindGt is True:
                    tax = (DecList[gwgzDex]+DecList[xjgzDex]+DecList[zbDex]+DecList[shbfDex]+DecList[xlfDex]+DecList[shfDex]+\
                        DecList[xnbtDex]+DecList[gtDex])/DecList[srzj3Dex]*DecList[ksh2Dex]
                elif isFindGt is False:
                    tax = (DecList[gwgzDex]+DecList[xjgzDex]+DecList[zbDex]+DecList[shbfDex]+DecList[xlfDex]+DecList[shfDex]+\
                        DecList[xnbtDex]+float(bzgt))/DecList[srzj3Dex]*DecList[ksh2Dex]
                ws2['P5'] = round(tax, 2)
                if isFindGt is True:
                    total = DecList[gwgzDex]+DecList[xjgzDex]+DecList[zbDex]+DecList[shbfDex]+DecList[xlfDex]+\
                        DecList[shfDex]+DecList[xnbtDex]+DecList[gtDex]-abs(DecList[jsheDex])-abs(DecList[kgjjDex])-abs(DecList[kzhynjDex])-\
                            abs(DecList[kyxDex])-abs(DecList[kshxDex])-abs(DecList[kylxDex])-tax
                elif isFindGt is False:
                    total = DecList[gwgzDex]+DecList[xjgzDex]+DecList[zbDex]+DecList[shbfDex]+DecList[xlfDex]+\
                        DecList[shfDex]+DecList[xnbtDex]+float(bzgt)-abs(DecList[jsheDex])-abs(DecList[kgjjDex])-abs(DecList[kzhynjDex])-\
                            abs(DecList[kyxDex])-abs(DecList[kshxDex])-abs(DecList[kylxDex])-tax
                ws2['Q5'] = round(total, 2)
                if total <= 3000:
                    partyFee = total*0.005
                elif total > 3000 and total <= 5000:
                    partyFee = total*0.01
                elif total > 5000 and total <= 10000:
                    partyFee = total*0.015
                elif total > 10000:
                    partyFee = total*0.02
                ws2['R5'] = round(partyFee, 2)
        except UnboundLocalError:
            QtWidgets.QMessageBox.information(self, "错误提示", "无论您需要计算第几季度的党费，本软件都需要您从财务系统导出工资数据表时，起始月份必须从【1月】开始，这也是财务"\
                                                    "系统的默认设置，所以您在从财务系统导出工资数据时，一定【不要】手动更改起始月份，点击查询工资后，再直接点击执行按钮导出工资表就可以啦，"\
                                                    "不需要做其他任何多余的操作。"\
                                                    "\n本程序还会自动读取电脑右下角的系统时间，根据系统时间，可以智能识别并读取工资表中当前季度所需的工资数据，并用于计算党费。"\
                                                    "\n很遗憾，本次生成党费缴费表失败！请您按要求重新到财务系统导出工资数据表后再试!"\
                                                    "\n\n点击OK后，程序将自动退出", QtWidgets.QMessageBox.Ok)
            exit()
        totalPartyFee = round(ws2['R3'].value + ws2['R4'].value + ws2['R5'].value, 2)
        ws2['B6'] = totalPartyFee
        ws2['E9'] = '应交总金额：'+str(totalPartyFee)+'元'
        ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
        ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
        # time.sleep(2)
        partyFeeName = str(year)+'年第'+str(season)+'季度_应缴党费.xlsx'
        wb2.save(current_dir + '/' + partyFeeName)
        # print()
        # print('正在生成应交党费数据...')
        # time.sleep(2)
        # input('工资数据写入成功！保存文件名为【'+partyFeeName+'】，按回车键退出程序！')
        QtWidgets.QMessageBox.information(self, "成功提示", "缴费表格已生成！\n生成的缴费表格文件所在路径和原始工资文件一致。\n文件名为：" + partyFeeName, QtWidgets.QMessageBox.Ok)
if __name__ == '__main__':
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"         #高分辨率支持
    app = QtWidgets.QApplication(sys.argv)
    app.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)            #高分辨率支持
    window = mywindow()
    window.show()
    sys.exit(app.exec_())

